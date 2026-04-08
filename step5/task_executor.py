"""
Step 5 - Task Executor
========================
Execute research tasks by:
  1. Reading the actual dataset file
  2. Sending the agent prompt (from agent_prompt.txt) + task details + data to GPT
  3. GPT generates analysis Python code
  4. Executing the generated code to produce plots and analysis files
"""

import os
import json
import logging
import subprocess
import sys
from typing import Dict, Any, List, Optional

from step5.gpt_client import call_gpt

logger = logging.getLogger(__name__)


# ===================================================================
# Minimal user message: task JSON + file paths only
# ===================================================================

TASK_USER_MSG = """Execute this task. The dataset file is at: {data_path}
Save figure to: {output_figure_path}
Save analysis to: {output_analysis_path}

Dataset preview:
{data_preview}

Task:
{task_json}

Output a single self-contained Python script. No markdown fences.
"""


def execute_task(
    task: Dict[str, Any],
    paper_dir: str,
    output_dir: str,
    agent_prompt: str,
    config: Dict[str, Any],
    model: str = "gpt-5.4-mini",
) -> Dict[str, Any]:
    """
    Execute a single research task using the agent prompt as system prompt.

    Args:
        task: Task dict from tasks.json
        paper_dir: Path to organized paper directory (with type1_data/, type2_data/)
        output_dir: Path to save outputs for this task
        agent_prompt: The agent prompt text (from agent_prompt.txt) — used as system prompt
        config: Step 5 config dict
        model: GPT model

    Returns:
        Dict with execution result
    """
    task_id = task.get("task_id", "UNKNOWN")
    os.makedirs(output_dir, exist_ok=True)

    # Find the dataset file
    dataset_file = task.get("required_inputs", {}).get("dataset_file", "")
    data_path = _find_dataset_file(dataset_file, paper_dir)

    if not data_path:
        msg = f"Dataset file not found: {dataset_file}"
        logger.warning(f"    {msg}")
        _save_error_report(output_dir, task_id, msg)
        return {"status": "error", "error": msg, "output_files": []}

    logger.info(f"    Dataset: {os.path.basename(data_path)} ({os.path.getsize(data_path)//1024}KB)")

    # Generate data preview
    data_preview = _generate_data_preview(data_path)

    # Define output paths
    fig_filename = f"{task_id}_figure.png"
    analysis_filename = f"{task_id}_Analysis.txt"
    script_filename = f"{task_id}_script.py"

    output_figure_path = os.path.join(output_dir, fig_filename)
    output_analysis_path = os.path.join(output_dir, analysis_filename)
    script_path = os.path.join(output_dir, script_filename)

    # Build user message: just task JSON + file paths
    user_prompt = TASK_USER_MSG.format(
        task_json=json.dumps(task, indent=2, ensure_ascii=False),
        data_preview=data_preview,
        data_path=data_path.replace("\\", "/"),
        output_figure_path=output_figure_path.replace("\\", "/"),
        output_analysis_path=output_analysis_path.replace("\\", "/"),
    )

    # Call GPT: agent_prompt as system, task details as user
    # Retry up to 2 times if the generated code fails, feeding error back
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    max_retries = 2
    last_error = ""

    for attempt in range(max_retries + 1):
        prompt_to_send = user_prompt
        if attempt > 0 and last_error:
            prompt_to_send += f"\n\nPREVIOUS ATTEMPT FAILED with error:\n{last_error}\nPlease fix the issue and regenerate the script."

        try:
            code = call_gpt(
                prompt=prompt_to_send,
                system_prompt=agent_prompt,
                model=model,
                temperature=0.2,
                max_tokens=6000,
            )
        except Exception as e:
            msg = f"Code generation GPT failed: {e}"
            logger.warning(f"    {msg}")
            _save_error_report(output_dir, task_id, msg)
            return {"status": "error", "error": msg, "output_files": []}

        # Clean code
        code = _clean_code(code)

        # Save the script
        with open(script_path, "w", encoding="utf-8") as f:
            f.write(code)
        logger.info(f"    Script saved: {script_filename} (attempt {attempt+1})")

        # Execute
        result = _execute_script(script_path, project_root)

        if result["success"]:
            break

        last_error = result.get("stderr", "")[:800]
        if attempt < max_retries:
            logger.info(f"    Retrying code generation (attempt {attempt+2})...")

    # Check outputs
    output_files = []
    if os.path.exists(output_figure_path):
        output_files.append(fig_filename)
        logger.info(f"    Figure saved: {fig_filename}")
    else:
        logger.warning(f"    Figure NOT generated: {fig_filename}")

    if os.path.exists(output_analysis_path):
        output_files.append(analysis_filename)
        logger.info(f"    Analysis saved: {analysis_filename}")
    else:
        logger.warning(f"    Analysis NOT generated: {analysis_filename}")

    output_files.append(script_filename)

    return {
        "status": "success" if result["success"] else "partial",
        "output_files": output_files,
        "has_figure": os.path.exists(output_figure_path),
        "has_analysis": os.path.exists(output_analysis_path),
        "exit_code": result.get("exit_code", -1),
        "stderr": result.get("stderr", "")[:500],
    }


def _find_dataset_file(filename: str, paper_dir: str) -> Optional[str]:
    """Search for a dataset file in the organized paper directory."""
    if not filename:
        return None

    search_dirs = [
        os.path.join(paper_dir, "type1_data"),
        os.path.join(paper_dir, "type2_data"),
        os.path.join(paper_dir, "scripts"),
        paper_dir,
    ]

    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        candidate = os.path.join(d, filename)
        if os.path.exists(candidate):
            return candidate

    # Fuzzy match
    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        for f in os.listdir(d):
            if f.lower() == filename.lower():
                return os.path.join(d, f)
            if os.path.splitext(f)[0].lower().startswith(os.path.splitext(filename)[0].lower()):
                return os.path.join(d, f)

    return None


def _generate_data_preview(data_path: str) -> str:
    """Generate a text preview of the data file for the GPT prompt."""
    ext = os.path.splitext(data_path)[1].lower()
    size = os.path.getsize(data_path)
    lines = [f"File: {os.path.basename(data_path)} ({size//1024}KB)"]

    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(data_path, read_only=True, data_only=True)
            lines.append(f"Sheets: {wb.sheetnames}")
            for sname in wb.sheetnames[:3]:
                ws = wb[sname]
                lines.append(f"\nSheet '{sname}': {ws.max_row} rows x {ws.max_column} cols")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 8:
                        break
                    row_vals = [str(v) if v is not None else "" for v in row[:20]]
                    lines.append(f"  Row {i}: {row_vals}")
            wb.close()

        elif ext == ".csv":
            with open(data_path, "r", encoding="utf-8", errors="replace") as f:
                for i, line in enumerate(f):
                    if i >= 10:
                        lines.append("  ...")
                        break
                    lines.append(f"  Row {i}: {line.strip()[:200]}")

        elif ext in (".npy", ".npz"):
            import numpy as np
            if ext == ".npy":
                arr = np.load(data_path, allow_pickle=True)
                lines.append(f"Shape: {arr.shape}, Dtype: {arr.dtype}")
                if arr.ndim <= 2 and arr.size < 100:
                    lines.append(f"Values: {arr}")
                else:
                    lines.append(f"First values: {arr.flat[:20]}")
            else:
                data = np.load(data_path, allow_pickle=True)
                lines.append(f"Keys: {list(data.keys())}")

        elif ext in (".txt", ".dat", ".tsv"):
            with open(data_path, "r", encoding="utf-8", errors="replace") as f:
                for i, line in enumerate(f):
                    if i >= 10:
                        lines.append("  ...")
                        break
                    lines.append(f"  Line {i}: {line.strip()[:200]}")
        else:
            lines.append(f"Binary file, format: {ext}")

    except Exception as e:
        lines.append(f"(Preview error: {e})")

    return "\n".join(lines)


def _clean_code(code: str) -> str:
    """Remove markdown fences and other non-code content."""
    code = code.strip()
    if code.startswith("```python"):
        code = code[len("```python"):].strip()
    elif code.startswith("```"):
        code = code[3:].strip()
    if code.endswith("```"):
        code = code[:-3].strip()
    return code


def _execute_script(script_path: str, working_dir: str) -> Dict[str, Any]:
    """Execute a Python script in a subprocess."""
    script_path = os.path.abspath(script_path)
    working_dir = os.path.abspath(working_dir)
    try:
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=120,
            cwd=working_dir,
        )

        if result.stdout:
            logger.info(f"    stdout: {result.stdout[:300]}")
        if result.returncode != 0:
            logger.warning(f"    Script exited with code {result.returncode}")
            if result.stderr:
                logger.warning(f"    stderr: {result.stderr[:500]}")

        return {
            "success": result.returncode == 0,
            "exit_code": result.returncode,
            "stdout": result.stdout[:2000],
            "stderr": result.stderr[:2000],
        }

    except subprocess.TimeoutExpired:
        logger.warning("    Script timed out (120s)")
        return {"success": False, "exit_code": -1, "stdout": "", "stderr": "Timeout"}
    except Exception as e:
        logger.warning(f"    Script execution failed: {e}")
        return {"success": False, "exit_code": -1, "stdout": "", "stderr": str(e)}


def _save_error_report(output_dir: str, task_id: str, error_msg: str):
    """Save an error report when a task can't be completed."""
    path = os.path.join(output_dir, f"{task_id}_Analysis.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"Task {task_id} - Error Report\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"The task could not be completed due to:\n{error_msg}\n")
