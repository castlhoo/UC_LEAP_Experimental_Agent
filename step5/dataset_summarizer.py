"""
Step 5 - Dataset Summarizer
==============================
Read organized dataset files and generate a structured summary via GPT.
This is the 1st GPT call in Step 5.
"""

import os
import logging
from typing import Dict, Any, List, Optional

import openpyxl
import numpy as np

from step5.gpt_client import call_gpt

logger = logging.getLogger(__name__)


# ===================================================================
# Dataset Summary Prompt (1st GPT call)
# ===================================================================

SUMMARY_SYSTEM = """You are an expert in scientific data analysis, especially for experimental physics datasets."""

SUMMARY_PROMPT = """I will provide you with a dataset (or its schema, metadata, and sample values).
Your task is to analyze it and generate a structured, high-level summary that helps downstream task generation.

Please follow this format strictly:

1. Dataset Overview
- Format (e.g., QCoDeS SQLite, CSV, etc.)
- Number of runs / experiments
- Approximate data volume
- Time range (if available)

2. Instrumentation Setup
- List all instruments inferred from the dataset (e.g., lock-in amplifiers, temperature controllers, magnets, SMUs)
- Briefly describe the role of each instrument

3. Information of each run / experiment, do the following for all the runs / experiments in sequence: (if individual run/experiment information is not available, just provide the following information for all the data provided altogether)

a. Measured Parameters
- Independent variables (e.g., magnetic field, time, gate voltage)
- Dependent variables (e.g., voltage, resistance, temperature)
- Include naming conventions from the dataset

b. Experimental Context (Inference)
- What kind of experiment is this likely to be?
- Possible physical phenomena (e.g., quantum Hall effect, magnetotransport, etc.)
- Reasoning based on variables and instruments

c. Notes for Analysis
- Any important patterns in the dataset structure
- Potential challenges or ambiguities
- Suggestions for how this dataset could be analyzed

Important:
- Be concise but informative
- Use bullet points
- Do NOT hallucinate unsupported claims
- Clearly distinguish between observed facts and inferred interpretations

=== PAPER CONTEXT ===
Title: {title}
Journal: {journal}
Paper Summary: {paper_summary}

=== DATASET FILES ===
{file_summaries}
"""


def generate_dataset_summary(
    paper: Dict[str, Any],
    paper_dir: str,
    paper_analysis: Optional[Dict[str, Any]],
    config: Dict[str, Any],
    model: str = "gpt-5.4-mini",
) -> str:
    """
    Generate a structured dataset summary for a paper's organized files.

    Args:
        paper: Paper entry from Step 4 manifest
        paper_dir: Absolute path to paper's organized directory
        paper_analysis: Paper analysis from Step 3 (summary, figures, etc.)
        config: Step 5 config dict
        model: GPT model to use

    Returns:
        Dataset summary as text string
    """
    summary_config = config.get("summary", {})
    max_files = summary_config.get("max_files", 30)
    preview_rows = summary_config.get("preview_rows", 10)
    max_columns = summary_config.get("max_columns", 30)

    # Collect file previews
    file_summaries = []

    # Type 1 data
    t1_dir = os.path.join(paper_dir, "type1_data")
    if os.path.isdir(t1_dir):
        for fname in sorted(os.listdir(t1_dir))[:max_files]:
            fpath = os.path.join(t1_dir, fname)
            preview = _preview_file(fpath, preview_rows, max_columns)
            file_summaries.append(f"[Type 1 - replot-ready] {fname}\n{preview}")

    # Type 2 data
    t2_dir = os.path.join(paper_dir, "type2_data")
    if os.path.isdir(t2_dir):
        for fname in sorted(os.listdir(t2_dir))[:max_files]:
            fpath = os.path.join(t2_dir, fname)
            preview = _preview_file(fpath, preview_rows, max_columns)
            file_summaries.append(f"[Type 2 - raw data] {fname}\n{preview}")

    # Scripts
    script_dir = os.path.join(paper_dir, "scripts")
    if os.path.isdir(script_dir):
        for fname in sorted(os.listdir(script_dir))[:5]:
            fpath = os.path.join(script_dir, fname)
            preview = _preview_script(fpath)
            file_summaries.append(f"[Script] {fname}\n{preview}")

    if not file_summaries:
        return "(No dataset files available for summary)"

    # Build prompt
    paper_summary = ""
    if paper_analysis and paper_analysis.get("summary"):
        paper_summary = paper_analysis["summary"]

    prompt = SUMMARY_PROMPT.format(
        title=paper.get("title", ""),
        journal=paper.get("journal", ""),
        paper_summary=paper_summary,
        file_summaries="\n\n".join(file_summaries),
    )

    # Call GPT
    max_tokens = config.get("gpt", {}).get("summary_max_tokens", 4000)
    try:
        result = call_gpt(
            prompt=prompt,
            system_prompt=SUMMARY_SYSTEM,
            model=model,
            temperature=0.2,
            max_tokens=max_tokens,
        )
        return result
    except Exception as e:
        logger.warning(f"  Dataset summary GPT failed: {e}")
        return f"(GPT summary failed: {e})"


def _preview_file(fpath: str, max_rows: int = 10, max_cols: int = 30) -> str:
    """Generate a text preview of a data file."""
    ext = os.path.splitext(fpath)[1].lower()
    size = os.path.getsize(fpath)
    size_str = f"{size/1024:.1f}KB" if size < 1024*1024 else f"{size/1024/1024:.1f}MB"

    try:
        if ext in (".xlsx", ".xls"):
            return _preview_excel(fpath, max_rows, max_cols, size_str)
        elif ext == ".csv":
            return _preview_csv(fpath, max_rows, max_cols, size_str)
        elif ext in (".txt", ".dat", ".tsv"):
            return _preview_text(fpath, max_rows, size_str)
        elif ext in (".npy", ".npz"):
            return _preview_numpy(fpath, size_str)
        elif ext in (".sxm", ".ibw", ".spe"):
            return f"  Binary instrument file | Size: {size_str} | Format: {ext}"
        elif ext == ".json":
            return _preview_json(fpath, size_str)
        else:
            return f"  Binary file | Size: {size_str} | Format: {ext}"
    except Exception as e:
        return f"  (Preview failed: {e}) | Size: {size_str}"


def _preview_excel(fpath: str, max_rows: int, max_cols: int, size_str: str) -> str:
    """Preview Excel file."""
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    lines = [f"  Format: Excel | Size: {size_str} | Sheets: {wb.sheetnames}"]

    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows + 1:
                break
            row_vals = [str(v) if v is not None else "" for v in row[:max_cols]]
            rows.append(row_vals)

        if rows:
            # First row as headers
            headers = rows[0]
            lines.append(f"  Sheet '{sheet_name}': {ws.max_row} rows x {ws.max_column} cols")
            lines.append(f"    Headers: {headers}")
            for row in rows[1:min(4, len(rows))]:
                lines.append(f"    Sample: {row}")

    wb.close()
    return "\n".join(lines)


def _preview_csv(fpath: str, max_rows: int, max_cols: int, size_str: str) -> str:
    """Preview CSV file."""
    lines = [f"  Format: CSV | Size: {size_str}"]
    with open(fpath, "r", encoding="utf-8", errors="replace") as f:
        for i, line in enumerate(f):
            if i >= max_rows + 1:
                break
            line = line.strip()
            if i == 0:
                lines.append(f"  Headers: {line}")
            elif i <= 3:
                lines.append(f"  Row {i}: {line[:200]}")

    return "\n".join(lines)


def _preview_text(fpath: str, max_rows: int, size_str: str) -> str:
    """Preview text/dat/tsv file."""
    lines = [f"  Format: Text | Size: {size_str}"]
    with open(fpath, "r", encoding="utf-8", errors="replace") as f:
        for i, line in enumerate(f):
            if i >= max_rows:
                break
            line = line.strip()
            if i == 0:
                lines.append(f"  First line: {line[:200]}")
            elif i <= 3:
                lines.append(f"  Line {i+1}: {line[:200]}")

    return "\n".join(lines)


def _preview_numpy(fpath: str, size_str: str) -> str:
    """Preview numpy file."""
    ext = os.path.splitext(fpath)[1].lower()
    if ext == ".npy":
        arr = np.load(fpath, allow_pickle=True)
        return f"  Format: NumPy array | Size: {size_str} | Shape: {arr.shape} | Dtype: {arr.dtype}"
    elif ext == ".npz":
        data = np.load(fpath, allow_pickle=True)
        keys = list(data.keys())
        info = {k: f"shape={data[k].shape}, dtype={data[k].dtype}" for k in keys[:10]}
        return f"  Format: NumPy archive | Size: {size_str} | Keys: {info}"
    return f"  Format: NumPy | Size: {size_str}"


def _preview_json(fpath: str, size_str: str) -> str:
    """Preview JSON file."""
    import json
    with open(fpath, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        keys = list(data.keys())[:10]
        return f"  Format: JSON | Size: {size_str} | Top keys: {keys}"
    elif isinstance(data, list):
        return f"  Format: JSON array | Size: {size_str} | Length: {len(data)}"
    return f"  Format: JSON | Size: {size_str}"


def _preview_script(fpath: str) -> str:
    """Preview first few lines of a script."""
    lines = [f"  Format: Script | Size: {os.path.getsize(fpath)/1024:.1f}KB"]
    with open(fpath, "r", encoding="utf-8", errors="replace") as f:
        for i, line in enumerate(f):
            if i >= 5:
                break
            lines.append(f"  L{i+1}: {line.rstrip()[:120]}")
    return "\n".join(lines)
